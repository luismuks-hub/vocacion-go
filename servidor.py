#!/usr/bin/env python3
"""
Vocacion GO — Servidor web v3.0
Funciona tanto en local como en Render.com
"""

import sys, os, json, io, traceback, mimetypes
from datetime import datetime
from http.server import HTTPServer, BaseHTTPRequestHandler

BASE_DIR = os.path.dirname(os.path.abspath(__file__))
PORT     = int(os.environ.get('PORT', 8000))
IS_RENDER = 'RENDER' in os.environ

# ── Auto-instalar dependencias en local ──────────────────────
if not IS_RENDER:
    import subprocess
    for pkg in ['reportlab', 'openpyxl']:
        try:
            __import__(pkg)
        except ImportError:
            print(f'Instalando {pkg}...')
            subprocess.check_call(
                [sys.executable, '-m', 'pip', 'install', pkg, '--quiet'],
                stdout=subprocess.DEVNULL, stderr=subprocess.DEVNULL)


# ── Helpers ──────────────────────────────────────────────────
def limpiar(texto):
    """Convierte texto a ASCII para Helvetica en reportlab.
    Reemplaza tildes y ñ por equivalentes ASCII en lugar de eliminarlos."""
    if not texto: return ''
    REEMPLAZOS = {
        'a\u0301':'a','e\u0301':'e','i\u0301':'i','o\u0301':'o','u\u0301':'u',
        '\u00e1':'a','\u00e9':'e','\u00ed':'i','\u00f3':'o','\u00fa':'u',
        '\u00c1':'A','\u00c9':'E','\u00cd':'I','\u00d3':'O','\u00da':'U',
        '\u00f1':'n','\u00d1':'N','\u00fc':'u','\u00dc':'U',
        '\u00e0':'a','\u00e8':'e','\u00ec':'i','\u00f2':'o','\u00f9':'u',
        '\u2013':'-','\u2014':'-','\u2018':"'",'\u2019':"'",
        '\u201c':'"','\u201d':'"','\u2026':'...','\u00b7':'.',
        '\u00ba':'o','\u00aa':'a','\u00bf':'?','\u00a1':'!',
    }
    resultado = []
    for c in str(texto):
        if c in REEMPLAZOS:
            resultado.append(REEMPLAZOS[c])
        elif ord(c) < 128:
            resultado.append(c)
    return ''.join(resultado)

def truncar(texto, n):
    t = limpiar(texto)
    return t[:n-3] + '...' if len(t) > n else t


# ── GENERADOR PDF ────────────────────────────────────────────
def generar_pdf(datos):
    from reportlab.lib.pagesizes import A4
    from reportlab.lib.units import mm
    from reportlab.pdfgen import canvas as rl_canvas
    from reportlab.lib.colors import HexColor

    u  = datos.get('usuario', {})
    r  = datos.get('resultado', {})
    pn = r.get('puntajesNorm', {})
    nd = r.get('nivelesDims', {})
    carreras = r.get('carrerasRecomendadas', [])

    W, H = A4
    buf = io.BytesIO()
    c = rl_canvas.Canvas(buf, pagesize=A4)

    BG=HexColor('#0D0D1A'); PU=HexColor('#8B00FF'); LI=HexColor('#B8E800')
    BL=HexColor('#00A8DD'); WH=HexColor('#FFFFFF');  GL=HexColor('#DCDCEB')
    GM=HexColor('#9090B0'); GD=HexColor('#505080');  GR=HexColor('#30DD70')

    DIM_C = {
        'DIM-01':HexColor('#8B00FF'),'DIM-02':HexColor('#22C55E'),
        'DIM-03':HexColor('#00A8DD'),'DIM-04':HexColor('#DD0088'),
        'DIM-05':HexColor('#FF6B00'),'DIM-06':HexColor('#7080A0'),
    }
    DIM_LBL = {
        'DIM-01':'Academico','DIM-02':'Tecnico','DIM-03':'Cientifico',
        'DIM-04':'Social','DIM-05':'Serv. Publico','DIM-06':'Disciplina'
    }
    CAT_C = {
        'UNIVERSITARIA':HexColor('#8B00FF'),'UNIVERSITARIA_SOCIAL':HexColor('#8B00FF'),
        'UNIVERSITARIA_CIENCIAS':HexColor('#00A8DD'),'UNIVERSITARIA_TECNICA':HexColor('#8B00FF'),
        'TECNICA':HexColor('#22C55E'),'FUERZAS_ARMADAS':HexColor('#FF6B00'),
        'POLICIAL':HexColor('#DD0088'),
    }

    # Cabecera
    c.setFillColor(HexColor('#150830')); c.rect(0,H-56*mm,W,56*mm,fill=1,stroke=0)
    c.setFillColor(PU); c.rect(0,H-3*mm,W,3*mm,fill=1,stroke=0)
    c.setFillColor(WH); c.setFont('Helvetica-Bold',22)
    c.drawString(18*mm,H-18*mm,'VOCACION GO')
    c.setFillColor(LI); c.setFont('Helvetica-Bold',9)
    c.drawString(18*mm,H-26*mm,'EL JUEGO MAS IMPORTANTE: TU CARRERA')
    c.setFillColor(GL); c.setFont('Helvetica',12)
    nom = truncar(f"{u.get('nombre','')} {u.get('apellido','')}",50).strip()
    c.drawString(18*mm,H-36*mm,f'Reporte: {nom}')
    c.setFillColor(GM); c.setFont('Helvetica',8)
    fecha = datetime.now().strftime('%d/%m/%Y')
    c.drawString(18*mm,H-44*mm,
        f"{fecha} | {u.get('edad','?')} anios | {limpiar(u.get('region','?'))} | {limpiar(u.get('grado','?'))}")
    c.setFillColor(PU); c.rect(0,H-57*mm,W,3*mm,fill=1,stroke=0)

    # Fondo
    c.setFillColor(BG); c.rect(0,0,W,H-57*mm,fill=1,stroke=0)
    y = H-68*mm

    # Resultado
    c.setFillColor(HexColor('#1E0840')); c.setStrokeColor(PU); c.setLineWidth(0.8)
    c.roundRect(18*mm,y-24*mm,W-36*mm,24*mm,4*mm,fill=1,stroke=1)
    cat = limpiar(r.get('categoriaLabel','')).upper()
    c.setFillColor(LI); c.setFont('Helvetica-Bold',14)
    c.drawCentredString(W/2,y-10*mm,cat)
    c.setFillColor(GM); c.setFont('Helvetica',8)
    c.drawCentredString(W/2,y-18*mm,
        f"Regla: {limpiar(r.get('reglaAplicada','?'))} | Confianza: {limpiar(r.get('confianza','?'))}")
    y -= 33*mm

    # Dimensiones
    c.setFillColor(BL); c.setFont('Helvetica-Bold',10)
    c.drawString(18*mm,y,'PERFIL POR DIMENSIONES')
    y -= 4*mm
    c.setStrokeColor(GD); c.setLineWidth(0.4); c.line(18*mm,y,W-18*mm,y)
    y -= 7*mm

    for i,did in enumerate(['DIM-01','DIM-02','DIM-03','DIM-04','DIM-05','DIM-06']):
        p = pn.get(did,0); cx=18*mm if i%2==0 else W/2+5*mm; ry=y-(i//2)*17*mm
        dc = DIM_C.get(did,GM)
        c.setFillColor(dc); c.setFont('Helvetica-Bold',8)
        c.drawString(cx,ry,DIM_LBL.get(did,did))
        c.setFillColor(WH); c.drawString(cx+48*mm,ry,f'{p}/100')
        c.setFillColor(GD); c.roundRect(cx,ry-5.5*mm,72*mm,3.5*mm,1*mm,fill=1,stroke=0)
        if p > 0:
            c.setFillColor(dc); c.roundRect(cx,ry-5.5*mm,72*mm*(p/100),3.5*mm,1*mm,fill=1,stroke=0)
        nv = limpiar(nd.get(did,{}).get('nivel',''))
        c.setFillColor(GM); c.setFont('Helvetica',7); c.drawString(cx+58*mm,ry,nv)

    y -= 3*17*mm + 8*mm

    # Mensaje
    msg = truncar(r.get('mensaje',''),130)
    if msg:
        c.setFillColor(HexColor('#14142A'))
        c.roundRect(18*mm,y-13*mm,W-36*mm,13*mm,3*mm,fill=1,stroke=0)
        c.setFillColor(GL); c.setFont('Helvetica-Oblique',8)
        c.drawCentredString(W/2,y-8*mm,msg); y -= 22*mm

    # Subcategorías con carreras individuales
    c.setFillColor(BL); c.setFont('Helvetica-Bold',10)
    c.drawString(18*mm,y,'CARRERAS RECOMENDADAS')
    y -= 4*mm
    c.setStrokeColor(GD); c.setLineWidth(0.3); c.line(18*mm,y,W-18*mm,y)
    y -= 7*mm

    pag = 1

    def nueva_pagina_pdf():
        nonlocal y, pag
        c.setFillColor(BG); c.rect(0,0,W,10*mm,fill=1,stroke=0)
        c.setFillColor(GD); c.setFont('Helvetica',6)
        c.drawString(18*mm,4*mm,'Vocacion GO | Resultado orientativo. No reemplaza orientacion vocacional presencial.')
        c.drawRightString(W-18*mm,4*mm,f'Pag. {pag}')
        c.setFillColor(PU); c.rect(0,0,W,2*mm,fill=1,stroke=0)
        c.showPage()
        c.setFillColor(BG); c.rect(0,0,W,H,fill=1,stroke=0)
        c.setFillColor(HexColor('#150830')); c.rect(0,H-14*mm,W,14*mm,fill=1,stroke=0)
        c.setFillColor(LI); c.setFont('Helvetica-Bold',8)
        c.drawString(18*mm,H-9*mm,'VOCACION GO - Carreras recomendadas (continuacion)')
        c.setFillColor(PU); c.rect(0,H-14*mm,W,1.5*mm,fill=1,stroke=0)
        y = H-22*mm; pag += 1

    for sc in carreras[:6]:
        if y < 30*mm: nueva_pagina_pdf()
        ac = CAT_C.get(sc.get('categoria',''), GM)

        # ── Header de subcategoría ──
        c.setFillColor(HexColor('#1A1030'))
        c.roundRect(18*mm, y-8*mm, W-36*mm, 8*mm, 2*mm, fill=1, stroke=0)
        c.setFillColor(ac); c.setFont('Helvetica-Bold',9)
        subcat = truncar(sc.get('subcategoria',''), 45)
        c.drawString(20*mm, y-5.5*mm, subcat)
        c.setFillColor(LI); c.setFont('Helvetica-Bold',9)
        c.drawRightString(W-20*mm, y-5.5*mm, f"{sc.get('afinidad',0)}% afin | {limpiar(sc.get('duracion',''))}")
        y -= 10*mm

        # ── Carreras individuales ──
        car_items = sc.get('carreras', [])
        # Si carreras es lista de strings (legado), convertir
        if car_items and isinstance(car_items[0], str):
            car_items = [{'nombre': n, 'icono': '', 'descripcion': '', 'campoLaboral': [], 'especializaciones': [], 'instituciones': sc.get('instituciones',[])} for n in car_items]

        for car in car_items:
            car_nombre = truncar(car.get('nombre',''), 55)
            car_desc   = truncar(car.get('descripcion',''), 200)
            campos     = [limpiar(x) for x in car.get('campoLaboral',[]) if limpiar(x)]
            especs     = [limpiar(x) for x in car.get('especializaciones',[]) if limpiar(x)]
            insts      = [limpiar(x) for x in car.get('instituciones',[]) if limpiar(x)]

            # Calcular altura del bloque
            altura = 7*mm  # nombre
            if car_desc: altura += 7*mm
            if campos:   altura += 5*mm
            if especs:   altura += 5*mm
            if insts:    altura += 5*mm
            altura += 3*mm  # padding inferior

            if y < 20*mm + altura: nueva_pagina_pdf()

            # Bullet y nombre de la carrera
            c.setFillColor(ac)
            c.circle(20*mm, y-3*mm, 1.5*mm, fill=1, stroke=0)
            c.setFont('Helvetica-Bold',9)
            c.drawString(23*mm, y-3*mm, car_nombre)
            y -= 7*mm

            # Descripción (puede ser larga — dividir en líneas de ~85 chars)
            if car_desc:
                c.setFillColor(GM); c.setFont('Helvetica-Oblique',7.5)
                # Dividir en líneas de 85 caracteres máximo
                palabras = car_desc.split()
                lineas_desc = []
                linea_actual = ''
                for palabra in palabras:
                    prueba = (linea_actual + ' ' + palabra).strip()
                    if len(prueba) <= 90:
                        linea_actual = prueba
                    else:
                        if linea_actual:
                            lineas_desc.append(linea_actual)
                        linea_actual = palabra
                if linea_actual:
                    lineas_desc.append(linea_actual)
                for linea in lineas_desc[:3]:  # máximo 3 líneas
                    if y < 25*mm: nueva_pagina_pdf()
                    c.drawString(24*mm, y-2*mm, linea)
                    y -= 4.5*mm
                y -= 1.5*mm

            # Campo laboral
            if campos:
                c.setFillColor(HexColor('#5050A0')); c.setFont('Helvetica-Bold',7)
                c.drawString(24*mm, y-2*mm, 'Campo laboral:')
                c.setFillColor(GL); c.setFont('Helvetica',7)
                c.drawString(50*mm, y-2*mm, (', '.join(campos[:4]))[:70])
                y -= 5*mm

            # Especializaciones
            if especs:
                c.setFillColor(HexColor('#30607A')); c.setFont('Helvetica-Bold',7)
                c.drawString(24*mm, y-2*mm, 'Especializaciones:')
                c.setFillColor(GL); c.setFont('Helvetica',7)
                c.drawString(54*mm, y-2*mm, (', '.join(especs[:4]))[:65])
                y -= 5*mm

            # Instituciones
            if insts:
                c.setFillColor(HexColor('#706030')); c.setFont('Helvetica-Bold',7)
                c.drawString(24*mm, y-2*mm, 'Instituciones:')
                c.setFillColor(GL); c.setFont('Helvetica',7)
                c.drawString(47*mm, y-2*mm, (', '.join(insts[:4]))[:70])
                y -= 5*mm

            y -= 3*mm  # espacio entre carreras

        # Separador entre subcategorías
        c.setStrokeColor(GD); c.setLineWidth(0.2)
        c.line(18*mm, y, W-18*mm, y)
        y -= 5*mm

    # Pie última página
    c.setFillColor(BG); c.rect(0,0,W,10*mm,fill=1,stroke=0)
    c.setFillColor(GD); c.setFont('Helvetica',6)
    c.drawString(18*mm,4*mm,'Vocacion GO | Resultado orientativo. No reemplaza orientacion vocacional presencial.')
    c.drawRightString(W-18*mm,4*mm,f'Pag. {pag}')
    c.setFillColor(PU); c.rect(0,0,W,2*mm,fill=1,stroke=0)
    c.save()
    return buf.getvalue()


# ── GENERADOR EXCEL ──────────────────────────────────────────
def generar_excel(datos):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    u=datos.get('usuario',{}); r=datos.get('resultado',{})
    pn=r.get('puntajesNorm',{}); pb=r.get('puntajesBrutos',{})
    nd=r.get('nivelesDims',{}); resp=datos.get('respuestas',{})
    items=datos.get('items',[]); todas=datos.get('todasEvaluaciones',[])

    wb = openpyxl.Workbook()

    def hfill(h): return PatternFill('solid',fgColor=h.lstrip('#'))

    DIMS=[('DIM-01','Academico',10,1.0),('DIM-02','Tecnico',10,1.0),
          ('DIM-03','Cientifico',10,1.0),('DIM-04','Social',10,1.0),
          ('DIM-05','Serv. Publico',8,1.25),('DIM-06','Disciplina',8,1.25)]
    DIM_NOM={d[0]:d[1] for d in DIMS}
    ETQ={1:'Totalmente en desacuerdo',2:'En desacuerdo',
         3:'Ni de acuerdo ni en desacuerdo',4:'De acuerdo',5:'Totalmente de acuerdo'}

    ws1=wb.active; ws1.title='Resumen'

    def wh(ws,row,txt,color='8B00FF',sz=12):
        cell=ws.cell(row=row,column=1,value=txt)
        cell.font=Font(bold=True,color=color,size=sz); cell.fill=hfill('#0D0D1A'); return row+1

    def wr(ws,row,k,v):
        ws.cell(row=row,column=1,value=k).font=Font(bold=True,color='6060A0')
        ws.cell(row=row,column=2,value=str(v) if v is not None else ''); return row+1

    row=1; row=wh(ws1,row,'VOCACION GO - Reporte','C8FF00',13); row+=1
    row=wh(ws1,row,'DATOS DEL EVALUADO','00C8FF',10)
    row=wr(ws1,row,'Nombre',f"{u.get('nombre','')} {u.get('apellido','')}".strip())
    row=wr(ws1,row,'Edad',u.get('edad',''))
    row=wr(ws1,row,'Genero','Masculino' if u.get('genero')=='M' else ('Femenino' if u.get('genero')=='F' else 'No indica'))
    row=wr(ws1,row,'Grado',u.get('grado',''))
    row=wr(ws1,row,'Region',u.get('region',''))
    row=wr(ws1,row,'Institucion',u.get('institucion',''))
    row=wr(ws1,row,'Fecha',datetime.now().strftime('%d/%m/%Y'))
    row+=1; row=wh(ws1,row,'RESULTADO VOCACIONAL','00C8FF',10)
    row=wr(ws1,row,'Categoria',r.get('categoriaLabel',''))
    row=wr(ws1,row,'Regla',r.get('reglaAplicada',''))
    row=wr(ws1,row,'Confianza',r.get('confianza',''))
    row=wr(ws1,row,'Mensaje',r.get('mensaje',''))
    row+=1; row=wh(ws1,row,'PUNTAJES','00C8FF',10)
    for ci,h in enumerate(['Dimension','Codigo','Puntaje','Bruto','Nivel'],1):
        cell=ws1.cell(row=row,column=ci,value=h)
        cell.font=Font(bold=True,color='FFFFFF'); cell.fill=hfill('#1A1A35')
    row+=1
    for did,dnom,ni,_ in DIMS:
        ws1.append([dnom,did,pn.get(did,0),pb.get(did,0),nd.get(did,{}).get('nivel','')])
        row+=1
    row+=1; row=wh(ws1,row,'SUBCATEGORIAS RECOMENDADAS','00C8FF',10)
    ws1.append(['#','Subcategoria','Categoria','Afinidad %','Duracion','Sueldo Junior','Sueldo Senior'])
    row+=1
    for i,sc in enumerate(r.get('carrerasRecomendadas',[]),1):
        ws1.append([i,sc.get('subcategoria',sc.get('nombre','')),sc.get('categoria',''),
                    sc.get('afinidad',0),sc.get('duracion',''),sc.get('sueldoJunior',''),sc.get('sueldoSenior','')])
    for col,w in zip('ABCDEFG',[30,20,14,12,14,18,18]):
        ws1.column_dimensions[col].width=w

    # Hoja 2: Respuestas
    ws2=wb.create_sheet('Respuestas 56 items')
    for ci,h in enumerate(['ID','Dimension','Cod','Pregunta','Resp','Etiqueta','Holland','Kuder'],1):
        cell=ws2.cell(row=1,column=ci,value=h)
        cell.font=Font(bold=True,color='FFFFFF'); cell.fill=hfill('#1A1A35')
    DF={'DIM-01':'EDE9FE','DIM-02':'D1FAE5','DIM-03':'DBEAFE',
        'DIM-04':'FCE7F3','DIM-05':'FEF3C7','DIM-06':'F3F4F6'}
    for ri,item in enumerate(items,2):
        v=resp.get(item.get('id','')); vi=int(v) if str(v).isdigit() else None
        row_data=[item.get('id',''),DIM_NOM.get(item.get('dim',''),item.get('dim','')),
                  item.get('dim',''),item.get('texto',''),vi,ETQ.get(vi,'Sin respuesta' if vi is None else ''),
                  item.get('holland',''),item.get('kuder','')]
        for ci,val in enumerate(row_data,1):
            cell=ws2.cell(row=ri,column=ci,value=val)
            cell.fill=hfill(DF.get(item.get('dim',''),'FFFFFF'))
    for col,w in zip('ABCDEFGH',[9,32,8,65,10,34,22,14]):
        ws2.column_dimensions[col].width=w
    ws2.freeze_panes='A2'

    # Hoja 3: Puntajes
    ws3=wb.create_sheet('Puntajes')
    ws3.cell(row=1,column=1,value='CALCULO DE PUNTAJES').font=Font(bold=True,color='00C8FF',size=11)
    ws3.append([])
    for ci,h in enumerate(['Dimension','Codigo','N Items','Peso','Bruto','Max','Normalizado','Nivel'],1):
        cell=ws3.cell(row=3,column=ci,value=h)
        cell.font=Font(bold=True,color='FFFFFF'); cell.fill=hfill('#1A1A35')
    for did,dnom,ni,peso in DIMS:
        ws3.append([dnom,did,ni,peso,pb.get(did,0),ni*5,pn.get(did,0),nd.get(did,{}).get('nivel','')])
    for col,w in zip('ABCDEFGH',[35,8,8,6,10,6,20,12]):
        ws3.column_dimensions[col].width=w

    # Hoja 4: Comparación
    if todas:
        ws4=wb.create_sheet('Comparacion')
        ws4.cell(row=1,column=1,value='COMPARACION').font=Font(bold=True,color='00C8FF',size=11)
        ws4.append([])
        for ci,h in enumerate(['Nombre','Edad','Region','Fecha','Categoria','Confianza','Regla',
                                'DIM-01','DIM-02','DIM-03','DIM-04','DIM-05','DIM-06'],1):
            cell=ws4.cell(row=3,column=ci,value=h)
            cell.font=Font(bold=True,color='FFFFFF'); cell.fill=hfill('#1A1A35')
        for ev in todas:
            eu=ev.get('usuario',{}); er=ev.get('resultado',{}); ep=er.get('puntajesNorm',{})
            fev=''
            try: fev=datetime.fromisoformat(ev.get('fechaCreacion','')).strftime('%d/%m/%Y')
            except: pass
            ws4.append([f"{eu.get('nombre','')} {eu.get('apellido','')}".strip(),
                        eu.get('edad',''),eu.get('region',''),fev,
                        er.get('categoriaLabel',''),er.get('confianza',''),er.get('reglaAplicada',''),
                        ep.get('DIM-01',''),ep.get('DIM-02',''),ep.get('DIM-03',''),
                        ep.get('DIM-04',''),ep.get('DIM-05',''),ep.get('DIM-06','')])
        for col,w in zip('ABCDEFGHIJKLM',[24,6,16,12,24,12,8,8,8,8,8,8,8]):
            ws4.column_dimensions[col].width=w
        ws4.freeze_panes='A4'

    buf=io.BytesIO(); wb.save(buf); return buf.getvalue()


# ── EXCEL TODAS LAS EVALUACIONES ─────────────────────────────
def generar_excel_todas(datos):
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment

    todas=datos.get('todasEvaluaciones',[]); items=datos.get('items',[])

    def hfill(h): return PatternFill('solid',fgColor=h.lstrip('#'))

    DIMS=[('DIM-01','Academico',10,1.0),('DIM-02','Tecnico',10,1.0),
          ('DIM-03','Cientifico',10,1.0),('DIM-04','Social',10,1.0),
          ('DIM-05','Serv. Publico',8,1.25),('DIM-06','Disciplina',8,1.25)]
    DIM_NOM={d[0]:d[1] for d in DIMS}
    ETQ={1:'Totalmente en desacuerdo',2:'En desacuerdo',
         3:'Ni de acuerdo ni en desacuerdo',4:'De acuerdo',5:'Totalmente de acuerdo'}
    DF={'DIM-01':'EDE9FE','DIM-02':'D1FAE5','DIM-03':'DBEAFE',
        'DIM-04':'FCE7F3','DIM-05':'FEF3C7','DIM-06':'F3F4F6'}

    wb=openpyxl.Workbook(); ws1=wb.active; ws1.title='Comparacion General'
    ws1.cell(row=1,column=1,value=f'VOCACION GO - Todas las Evaluaciones ({len(todas)})').font=Font(bold=True,color='C8FF00',size=13)
    ws1.append([])
    cab=['N','Nombre','Edad','Genero','Region','Grado','Fecha',
         'Categoria','Regla','Confianza',
         'DIM-01','DIM-02','DIM-03','DIM-04','DIM-05','DIM-06']
    for ci,h in enumerate(cab,1):
        cell=ws1.cell(row=3,column=ci,value=h)
        cell.font=Font(bold=True,color='FFFFFF'); cell.fill=hfill('#1A1A35')
        cell.alignment=Alignment(horizontal='center',wrap_text=True)
    for idx,ev in enumerate(todas,1):
        eu=ev.get('usuario',{}); er=ev.get('resultado',{}); ep=er.get('puntajesNorm',{})
        fev=''
        try: fev=datetime.fromisoformat(ev.get('fechaCreacion','')).strftime('%d/%m/%Y')
        except: pass
        gen='Masculino' if eu.get('genero')=='M' else ('Femenino' if eu.get('genero')=='F' else 'No indica')
        ws1.append([idx,f"{eu.get('nombre','')} {eu.get('apellido','')}".strip(),
                    eu.get('edad',''),gen,eu.get('region',''),eu.get('grado',''),fev,
                    er.get('categoriaLabel',''),er.get('reglaAplicada',''),er.get('confianza',''),
                    ep.get('DIM-01',''),ep.get('DIM-02',''),ep.get('DIM-03',''),
                    ep.get('DIM-04',''),ep.get('DIM-05',''),ep.get('DIM-06','')])
    for col,w in zip('ABCDEFGHIJKLMNOP',[4,24,6,12,16,18,12,24,8,12,8,8,8,8,8,8]):
        ws1.column_dimensions[col].width=w
    ws1.freeze_panes='A4'

    for idx,ev in enumerate(todas,1):
        eu=ev.get('usuario',{}); er=ev.get('resultado',{}); ep=er.get('puntajesNorm',{})
        pn=ep; pb=er.get('puntajesBrutos',{}); nd=er.get('nivelesDims',{})
        resp_ev=ev.get('respuestas',{})
        hoja=f"{idx}_{eu.get('nombre','?')[:8]}_{eu.get('apellido','?')[:6]}"[:31]
        ws=wb.create_sheet(hoja)
        ws.cell(row=1,column=1,value=f"Eval #{idx}: {eu.get('nombre','')} {eu.get('apellido','')}").font=Font(bold=True,color='C8FF00',size=11)
        ws.cell(row=2,column=1,value=f"Cat: {er.get('categoriaLabel','')} | Conf: {er.get('confianza','')} | Regla: {er.get('reglaAplicada','')}").font=Font(color='9090B0')
        ws.append([])
        ws.cell(row=4,column=1,value='PUNTAJES').font=Font(bold=True,color='00C8FF')
        for ci,h in enumerate(['Dim','Codigo','Puntaje','Nivel'],1):
            cell=ws.cell(row=5,column=ci,value=h); cell.font=Font(bold=True,color='FFFFFF'); cell.fill=hfill('#1A1A35')
        for did,dnom,_,_ in DIMS:
            ws.append([dnom,did,pn.get(did,''),nd.get(did,{}).get('nivel','')])
        ws.append([])
        r_row=ws.max_row+1
        ws.cell(row=r_row,column=1,value='RESPUESTAS').font=Font(bold=True,color='00C8FF')
        r_row+=1
        for ci,h in enumerate(['ID','Dimension','Pregunta','Resp','Etiqueta'],1):
            cell=ws.cell(row=r_row,column=ci,value=h); cell.font=Font(bold=True,color='FFFFFF'); cell.fill=hfill('#1A1A35')
        r_row+=1
        for item in items:
            v=resp_ev.get(item.get('id','')); vi=int(v) if str(v).isdigit() else None
            for ci,val in enumerate([item.get('id',''),DIM_NOM.get(item.get('dim',''),item.get('dim','')),
                                     item.get('texto',''),vi,ETQ.get(vi,'Sin respuesta' if vi is None else '')],1):
                cell=ws.cell(row=r_row,column=ci,value=val)
                cell.fill=hfill(DF.get(item.get('dim',''),'FFFFFF'))
            r_row+=1
        for col,w in zip('ABCDE',[9,16,65,10,34]):
            ws.column_dimensions[col].width=w
        ws.freeze_panes='A2'

    buf=io.BytesIO(); wb.save(buf); return buf.getvalue()


# ── SERVIDOR HTTP ─────────────────────────────────────────────
class Handler(BaseHTTPRequestHandler):

    def log_message(self, fmt, *args):
        # Mostrar solo errores (no 200/304)
        if args and str(args[1]) not in ('200','304','206'):
            print(f'[{args[1]}] {self.path}')

    def do_OPTIONS(self):
        self.send_response(200)
        self._cors()
        self.end_headers()

    def _cors(self):
        self.send_header('Access-Control-Allow-Origin', '*')
        self.send_header('Access-Control-Allow-Methods', 'GET,POST,OPTIONS')
        self.send_header('Access-Control-Allow-Headers', 'Content-Type')

    def do_GET(self):
        # Servir archivos estáticos
        path = self.path.split('?')[0]  # ignorar query strings

        # Ruta raíz → index.html
        if path == '/' or path == '':
            path = '/index.html'

        # Construir ruta absoluta al archivo
        file_path = os.path.join(BASE_DIR, path.lstrip('/'))

        # Seguridad: no salir del BASE_DIR
        file_path = os.path.realpath(file_path)
        if not file_path.startswith(os.path.realpath(BASE_DIR)):
            self.send_error(403, 'Acceso denegado')
            return

        if not os.path.isfile(file_path):
            self.send_error(404, f'Archivo no encontrado: {path}')
            return

        # Determinar Content-Type
        mime, _ = mimetypes.guess_type(file_path)
        if not mime:
            mime = 'application/octet-stream'

        # Leer y enviar
        try:
            with open(file_path, 'rb') as f:
                content = f.read()
            self.send_response(200)
            self.send_header('Content-Type', mime)
            self.send_header('Content-Length', str(len(content)))
            self._cors()
            self.end_headers()
            self.wfile.write(content)
        except Exception as e:
            self.send_error(500, str(e))

    def do_POST(self):
        if self.path == '/api/pdf':
            self._exportar(generar_pdf, 'application/pdf', 'VocacionGO.pdf')
        elif self.path == '/api/excel':
            ct = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            self._exportar(generar_excel, ct, 'VocacionGO.xlsx')
        elif self.path == '/api/excel/todas':
            ct = 'application/vnd.openxmlformats-officedocument.spreadsheetml.sheet'
            self._exportar(generar_excel_todas, ct, 'VocacionGO_Todas.xlsx')
        else:
            self.send_error(404, 'Endpoint no encontrado')

    def _exportar(self, fn, content_type, filename):
        try:
            length = int(self.headers.get('Content-Length', 0))
            body   = self.rfile.read(length)
            datos  = json.loads(body)
            archivo = fn(datos)
            self.send_response(200)
            self.send_header('Content-Type', content_type)
            self.send_header('Content-Disposition', f'attachment; filename="{filename}"')
            self.send_header('Content-Length', str(len(archivo)))
            self._cors()
            self.end_headers()
            self.wfile.write(archivo)
        except Exception as e:
            print(f'\nERROR exportar: {e}')
            traceback.print_exc()
            self.send_response(500)
            self.send_header('Content-Type', 'text/plain; charset=utf-8')
            self._cors()
            self.end_headers()
            self.wfile.write(str(e).encode('utf-8'))


# ── MAIN ──────────────────────────────────────────────────────
def main():
    os.chdir(BASE_DIR)
    port = int(os.environ.get('PORT', PORT))

    if IS_RENDER:
        print(f'VocacionGO arrancando en Render — puerto {port}')
        print(f'BASE_DIR: {BASE_DIR}')
        print(f'Archivos: {os.listdir(BASE_DIR)}')
    else:
        print('=' * 50)
        print('  VOCACION GO — Servidor local')
        print(f'  URL: http://localhost:{port}')
        print('  Ctrl+C para cerrar')
        print('=' * 50)

    HTTPServer.allow_reuse_address = True
    with HTTPServer(('0.0.0.0', port), Handler) as httpd:
        try:
            httpd.serve_forever()
        except KeyboardInterrupt:
            print('\nServidor cerrado.')

if __name__ == '__main__':
    main()
